"""
Excel Parser module for populating MongoDB with spreadsheet data.

This module provides functionality to parse Excel files and populate
the MongoDB database using the defined data models.
"""

import os
import logging
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils import range_boundaries

from db.models import Spreadsheet, Cell
from db.database import connect_db
from parsers.formula_parser import expand_cell_range

# Configure logging
logger = logging.getLogger(__name__)

class ExcelParser:
    """
    Parser for Excel spreadsheets that populates MongoDB using the Spreadsheet model.
    
    This class handles the extraction of data from Excel files and creates the
    appropriate document structures for storage in MongoDB.
    
    Attributes:
        file_path: Path to the Excel file.
        spreadsheet: The Spreadsheet document being populated.
    """
    
    def __init__(self, file_path: str):
        """
        Initialize the Excel parser with a file path.
        
        Args:
            file_path: Path to the Excel file to parse.
        """
        self.file_path = file_path
        self.spreadsheet = None
        self.workbook = None
        self.alias_mapping = {}
        self.reverse_alias_mapping = {}  # Maps aliases to cell references
    
    def parse(self) -> Spreadsheet:
        """
        Parse the Excel file and populate a Spreadsheet document.
        
        Returns:
            The populated Spreadsheet document.
        """
        logger.info(f"Parsing Excel file: {self.file_path}")
        
        # Check if this spreadsheet already exists in the database
        filename = os.path.basename(self.file_path)
        existing_spreadsheet = Spreadsheet.objects(file_path=self.file_path).first()
        
        if existing_spreadsheet:
            logger.info(f"Found existing spreadsheet: {existing_spreadsheet.name}")
            self.spreadsheet = existing_spreadsheet
            # Clear existing cells to avoid duplicates
            self.spreadsheet.cells = []
        else:
            # Create a new Spreadsheet document
            self.spreadsheet = Spreadsheet(
                name=Path(filename).stem,
                original_filename=filename,
                file_path=self.file_path
            )
        
        # Open the workbook
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=False)
        
        # Build alias mapping from defined names
        self._build_alias_mapping()
        
        # Get sheet names and set active sheet
        self.spreadsheet.sheet_names = self.workbook.sheetnames
        self.spreadsheet.active_sheet = self.workbook.active.title
        
        # Process each sheet
        for sheet_name in self.workbook.sheetnames:
            self._process_sheet(sheet_name)
        
        # Log summary of cells processed
        total_cells = len(self.spreadsheet.cells)
        logger.info(f"Total cells processed: {total_cells} from {len(self.workbook.sheetnames)} sheets")
        
        # Save the populated spreadsheet to the database first
        self.spreadsheet.save()
        
        # Process cell dependencies after all cells are saved
        self._process_cell_dependencies()
        
        # Save again after processing dependencies
        self.spreadsheet.save()
        
        logger.info(f"Completed parsing: {self.file_path}")
        return self.spreadsheet
    
    def _process_sheet(self, sheet_name: str) -> None:
        """
        Process a single sheet in the workbook.
        
        Args:
            sheet_name: Name of the sheet to process.
        """
        logger.info(f"Processing sheet: {sheet_name}")
        sheet = self.workbook[sheet_name]
        
        # Add sheet metadata to the spreadsheet document
        if 'sheets_metadata' not in self.spreadsheet.metadata:
            self.spreadsheet.metadata['sheets_metadata'] = {}
        
        self.spreadsheet.metadata['sheets_metadata'][sheet_name] = {
            'max_row': sheet.max_row,
            'max_column': sheet.max_column
        }
        
        logger.info(f"Sheet {sheet_name} dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
        
        # Process each non-empty cell
        cells_processed = 0
        for row_idx in range(1, sheet.max_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                # Skip empty cells
                if cell.value is None:
                    continue
                
                # Create and add the cell document
                cell_doc = self._create_cell_document(cell, row_idx-1, col_idx-1, sheet_name)
                self.spreadsheet.cells.append(cell_doc)
                cells_processed += 1
                
        logger.info(f"Processed {cells_processed} non-empty cells in sheet {sheet_name}")

    def get_cell_type(self, cell: openpyxl.cell.Cell) -> str:
        """
        Determine the type of cell content.
        - 'formula' if the cell contains a formula.
        - 'valuelist' if the cell has data validation with a list of values.
        - 'value' otherwise.
        """
        if cell.value is None:
            return None  # Skip empty cells.
        if isinstance(cell.value, str) and cell.value.startswith('='):
            return 'formula'
        
        # Check if the cell has data validation with a list
        validation_options = self._get_validation_options(cell)
        if validation_options:  # Now a list of dictionaries, but empty list still means no validation
            return 'valuelist'
        elif isinstance(cell.value, list):
            return 'valuelist'
        else:
            return 'value'
    
    def _get_validation_options(self, cell: openpyxl.cell.Cell) -> List[Dict[str, str]]:
        """
        Extract data validation options for a cell if it has list-type validation.
        
        Args:
            cell: The openpyxl Cell object.
            
        Returns:
            List of validation option dictionaries if the cell has list validation, empty list otherwise.
            Each dictionary contains workbook_name, sheet_name, and cell_ref.
        """
        worksheet = cell.parent
        cell_address = cell.coordinate
        workbook_name = os.path.basename(self.file_path)
        current_sheet_name = worksheet.title
        
        for dv in worksheet.data_validations.dataValidation:
            # Check if dv.sqref has 'ranges' (MultiCellRange) or is a string.
            if hasattr(dv.sqref, 'ranges'):
                ranges = dv.sqref.ranges
            else:
                ranges = dv.sqref.split()
            
            for cell_range in ranges:
                if self._cell_in_range(cell_address, cell_range):
                    if dv.type == "list":
                        # The options are stored in dv.formula1, e.g. '"Option1,Option2,Option3"'
                        formula = dv.formula1
                        # Remove surrounding quotes if present
                        if formula.startswith('"') and formula.endswith('"'):
                            formula = formula[1:-1]
                        options = formula.split(',')
                        expanded_options = []
                        for option in options:
                            option = option.strip()
                            if '!' in option:
                                sheet_name_option, range_part = option.split('!', 1)
                                cell_refs = expand_cell_range(range_part)
                                for cell_ref in cell_refs:
                                    clean_cell_ref = cell_ref.replace('$', '')
                                    expanded_options.append({
                                        "cell_ref": clean_cell_ref,
                                        "sheet_name": sheet_name_option,
                                        "workbook_name": workbook_name
                                    })
                            else:
                                cell_refs = expand_cell_range(option)
                                for cell_ref in cell_refs:
                                    clean_cell_ref = cell_ref.replace('$', '')
                                    expanded_options.append({
                                        "cell_ref": clean_cell_ref,
                                        "sheet_name": current_sheet_name,
                                        "workbook_name": workbook_name
                                    })
                        logger.info(f"Expanded options: {expanded_options}")
                        return expanded_options
        return []
    
    def _cell_in_range(self, cell, cell_range):
        """
        Check if a given cell (e.g. 'B2') is within the cell_range.
        cell_range can be a string or a CellRange object.
        
        Args:
            cell: Cell coordinate as string (e.g., 'B2')
            cell_range: Range to check against (string or CellRange object)
            
        Returns:
            True if the cell is in the range, False otherwise
        """
        col_letter, row = coordinate_from_string(cell)
        col = column_index_from_string(col_letter)
        
        if isinstance(cell_range, str):
            # If it's a string, use range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        else:
            # Otherwise, assume it's a CellRange object with min/max attributes
            min_col, min_row = cell_range.min_col, cell_range.min_row
            max_col, max_row = cell_range.max_col, cell_range.max_row

        return min_col <= col <= max_col and min_row <= row <= max_row
    
    def _build_alias_mapping(self) -> None:
        """
        Build a mapping of cell references to their aliases (defined names)
        and a reverse mapping from aliases to cell references.
        """
        logger.info("Building alias mapping from defined names")
        self.alias_mapping = {}
        self.reverse_alias_mapping = {}  # Maps aliases to cell references
        
        for alias, defined_name in self.workbook.defined_names.items():
            for sheet_name, cell_range in defined_name.destinations:
                cell_range_clean = cell_range.replace('$', '')
                if ':' in cell_range_clean:
                    continue  # skip ranges
                
                # Add to forward mapping (cell_ref -> alias)
                self.alias_mapping.setdefault(sheet_name, {})[cell_range_clean] = alias
                
                # Add to reverse mapping (alias -> cell_ref)
                self.reverse_alias_mapping[alias] = {
                    'sheet_name': sheet_name,
                    'cell_ref': cell_range_clean
                }
        
        total_aliases = sum(len(sheet_aliases) for sheet_aliases in self.alias_mapping.values())
        logger.info(f"Found {total_aliases} cell aliases across {len(self.alias_mapping)} sheets")

    def _create_cell_document(self, excel_cell: openpyxl.cell.Cell, 
                              row: int, column: int, sheet_name: str) -> Cell:
        """
        Create a Cell document from an Excel cell.
        
        Args:
            excel_cell: The openpyxl Cell object.
            row: The 0-indexed row number.
            column: The 0-indexed column number.
            sheet_name: The name of the sheet.
            
        Returns:
            A populated Cell document.
        """
        cell_ref = f"{get_column_letter(column+1)}{row+1}"
        
        # Determine cell type and data type
        
        data_type = type(excel_cell.value).__name__
        
        # Extract formula if present
        formula = None
        
        cell_type = self.get_cell_type(excel_cell)
        
        if cell_type == "formula":
            formula = excel_cell.value
        
        # Create value dictionary - flexible to store different types
        value = {
            "raw": excel_cell.value,
            "type": data_type
        }
        
        # Extract validation options if this is a valuelist cell
        validation_options = []
        if cell_type == "valuelist":
            validation_options = self._get_validation_options(excel_cell)
        
        # Check if this cell has an alias
        alias = None
        if sheet_name in self.alias_mapping:
            alias = self.alias_mapping[sheet_name].get(cell_ref)
        
        # Log cell information
        logger.debug(f"Cell {sheet_name}!{cell_ref}: type={cell_type}, data_type={data_type}, value={excel_cell.value}, alias={alias}")
        
        # Create the Cell document
        cell_doc = Cell(
            row=row,
            column=column,
            cell_reference=cell_ref,
            value=value,
            formatted_value=str(excel_cell.value),
            alias=alias,  # Set the alias from our mapping
            formula=formula,  # Use the extracted formula string
            sheet_name=sheet_name,  # Set sheet_name directly on the Cell
            data_type=data_type,
            cell_type=cell_type,
            value_list=validation_options,  # Store validation options in the value_list field
            metadata={
                "style": self._extract_cell_style(excel_cell)
            }
        )
        
        return cell_doc
    
    def _process_cell_dependencies(self) -> None:
        """
        Process all cells in the spreadsheet to establish precedents and dependents.
        This should be called after all cells have been created and saved to the database.
        """
        logger.info("Processing cell dependencies...")
        
        # Import the formula parser function
        from parsers.formula_parser import update_cell_dependencies
        
        # Get the workbook name from the file path
        workbook_name = os.path.basename(self.file_path)
        
        # Process each cell with a formula
        cells_with_formulas = [cell for cell in self.spreadsheet.cells if cell.formula]
        logger.info(f"Found {len(cells_with_formulas)} cells with formulas")
        
        # Clear all existing precedents and dependents first
        for cell in self.spreadsheet.cells:
            cell.precedent_cells = []
            cell.dependent_cells = []
            cell.formula_inputs = []
        
        # Now process each cell with a formula
        for cell in cells_with_formulas:
            # Ensure cell has a sheet_name
            if not cell.sheet_name:
                cell.sheet_name = self.spreadsheet.active_sheet
                
            # Update dependencies for this cell, passing the workbook name and alias mapping
            update_cell_dependencies(
                self.spreadsheet, 
                cell, 
                workbook_name, 
                reverse_alias_mapping=self.reverse_alias_mapping
            )
            
        logger.info("Cell dependency processing complete")
    
    def _extract_cell_style(self, excel_cell: openpyxl.cell.Cell) -> Dict[str, Any]:
        """
        Extract style information from an Excel cell.
        
        Args:
            excel_cell: The openpyxl Cell object.
            
        Returns:
            Dictionary containing style information.
        """
        style = {
            "number_format": excel_cell.number_format,
            "font": {
                "name": excel_cell.font.name,
                "size": excel_cell.font.size,
                "bold": excel_cell.font.bold,
                "italic": excel_cell.font.italic
            },
            "fill": {
                "type": excel_cell.fill.fill_type,
                "bg_color": excel_cell.fill.start_color.rgb if excel_cell.fill.start_color else None
            },
            "alignment": {
                "horizontal": excel_cell.alignment.horizontal,
                "vertical": excel_cell.alignment.vertical
            }
        }
        return style

def populate_database_from_excel(file_path: str) -> Spreadsheet:
    """
    Convenience function to parse an Excel file and populate the database.
    
    Args:
        file_path: Path to the Excel file.
        
    Returns:
        The populated Spreadsheet document.
    """
    parser = ExcelParser(file_path)
    return parser.parse()

def fix_missing_sheet_names(spreadsheet_name: str = None) -> int:
    """
    Fix any cells in the database that are missing the sheet_name field.
    
    This is a migration helper for spreadsheets that were created before
    the sheet_name field was added directly to the Cell model.
    
    Args:
        spreadsheet_name: Optional name of a specific spreadsheet to fix.
                          If None, fixes all spreadsheets.
    
    Returns:
        Number of cells that were fixed.
    """
    from db.models import Spreadsheet
    
    # Connect to the database if needed
    try:
        from db.database import connect_db
        connect_db()
    except Exception:
        # Assume we're already connected
        pass
    
    fixed_count = 0
    
    # Query for spreadsheets
    if spreadsheet_name:
        spreadsheets = Spreadsheet.objects(name=spreadsheet_name)
    else:
        spreadsheets = Spreadsheet.objects()
    
    # Process each spreadsheet
    for spreadsheet in spreadsheets:
        logger.info(f"Processing spreadsheet: {spreadsheet.name}")
        
        # Default to the active sheet if it exists
        default_sheet = spreadsheet.active_sheet or (
            spreadsheet.sheet_names[0] if spreadsheet.sheet_names else "Sheet1"
        )
        
        # Check for sheet_name in metadata
        if 'sheets_metadata' in spreadsheet.metadata:
            sheet_names = list(spreadsheet.metadata['sheets_metadata'].keys())
            
            # Process each cell
            for cell in spreadsheet.cells:
                if not cell.sheet_name:
                    # Try to determine sheet from metadata
                    sheet_in_metadata = None
                    if 'metadata' in cell and 'sheet_name' in cell.metadata:
                        sheet_in_metadata = cell.metadata['sheet_name']
                    
                    # Set the sheet_name
                    if sheet_in_metadata and sheet_in_metadata in sheet_names:
                        cell.sheet_name = sheet_in_metadata
                    else:
                        cell.sheet_name = default_sheet
                        
                    fixed_count += 1
        
        # Save the spreadsheet if any cells were fixed
        if fixed_count > 0:
            spreadsheet.save()
            logger.info(f"Fixed {fixed_count} cells in {spreadsheet.name}")
    
    return fixed_count

if __name__ == "__main__":
    import sys
    import argparse
    
    # Configure argument parser
    parser = argparse.ArgumentParser(description='Fix missing sheet_name fields in the database')
    parser.add_argument('--spreadsheet', help='Name of a specific spreadsheet to fix')
    args = parser.parse_args()
    
    # Run the fix
    count = fix_missing_sheet_names(args.spreadsheet)
    print(f"Fixed {count} cells")
    sys.exit(0) 